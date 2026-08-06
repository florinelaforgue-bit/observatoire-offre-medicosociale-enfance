"""
Module générique de manipulation et d'export de fichiers Excel (.xlsx) via openpyxl.
Aucune logique métier ni connaissance des données FINESS.
Compatible Python 3.13 et Termux Android.
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def creer_classeur() -> openpyxl.Workbook:
    """
    Crée et retourne un nouveau classeur Excel (Workbook)
    en supprimant la feuille par défaut si nécessaire pour repartir proprement.
    """
    wb = openpyxl.Workbook()
    # openpyxl crée une feuille par défaut nommée "Sheet"
    ws = wb.active
    ws.title = "Feuille1"
    return wb


def ajouter_feuille(wb: openpyxl.Workbook, nom_feuille: str) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Ajoute une nouvelle feuille au classeur avec le nom spécifié
    et retourne l'objet Worksheet correspondant.
    """
    # Évite les doublons de noms si la feuille existe déjà
    if nom_feuille in wb.sheetnames:
        # Suffixe pour éviter l'erreur d'unicité
        base_name = nom_feuille[:28]
        i = 1
        while f"{base_name}_{i}" in wb.sheetnames:
            i += 1
        nom_feuille = f"{base_name}_{i}"
        
    return wb.create_sheet(title=nom_feuille)


def ecrire_tableau(ws: openpyxl.worksheet.worksheet.Worksheet, donnees: list[dict], colonnes: list[str] = None) -> None:
    """
    Écrit un tableau de dictionnaires dans une feuille Excel de manière générique.
    - Ajoute les en-têtes en gras avec un fond coloré stylisé.
    - Écrit chaque ligne de données.
    - Fige la première ligne.
    - Active le filtre automatique.
    - Ajuste la largeur des colonnes.
    """
    if not donnees:
        return

    # Détermination des colonnes si non fournies explicitement
    if not colonnes:
        # Récupère l'ensemble des clés uniques de tous les dictionnaires pour préserver l'ordre d'apparition
        colonnes_set = []
        for ligne in donnees:
            for k in ligne.keys():
                if k not in colonnes_set:
                    colonnes_set.append(k)
        colonnes = colonnes_set

    # Styles professionnels (palette sobre et élégante, ex: gris anthracite / en-tête sombre)
    font_entete = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_entete = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    align_entete = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_fin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # 1. Écriture des en-têtes (Ligne 1)
    ws.append(colonnes)
    for col_num in range(1, len(colonnes) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_entete
        cell.fill = fill_entete
        cell.alignment = align_entete
        cell.border = border_fin

    # Hauteur de la ligne d'en-tête
    ws.row_dimensions[1].height = 25

    # 2. Écriture des lignes de données
    for ligne_idx, item in enumerate(donnees, start=2):
        valeurs_ligne = [item.get(col, "") for col in colonnes]
        ws.append(valeurs_ligne)
        
        # Application d'un style de bordure léger et d'un alignement par défaut
        for col_num in range(1, len(colonnes) + 1):
            cell = ws.cell(row=ligne_idx, column=col_num)
            cell.border = border_fin
            cell.alignment = Alignment(vertical="center")
            
        ws.row_dimensions[ligne_idx].height = 20

    # 3. Fonctionnalités de confort (Figer les volets et filtre automatique)
    ws.freeze_panes = "A2"
    
    derniere_col_lettre = get_column_letter(len(colonnes))
    derniere_ligne = len(donnees) + 1
    ws.auto_filter.ref = f"A1:{derniere_col_lettre}{derniere_ligne}"

    # 4. Ajustement automatique de la largeur des colonnes
    ajuster_colonnes(ws)


def ajuster_colonnes(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Ajuste dynamiquement la largeur des colonnes d'une feuille Excel
    en fonction de la longueur de leur contenu (avec une marge de sécurité).
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            valeur = str(cell.value) if cell.value is not None else ""
            if len(valeur) > max_len:
                max_len = len(valeur)
                
        # Application de la largeur calculée avec un minimum et un plafond raisonnable
        nouvelle_largeur = max(max_len + 3, 12)
        # Plafond pour éviter des colonnes démesurées (ex: textes longs)
        if nouvelle_largeur > 60:
            nouvelle_largeur = 60
            
        ws.column_dimensions[col_letter].width = nouvelle_largeur


def sauvegarder(wb: openpyxl.Workbook, chemin_destination: str | Path) -> None:
    """
    Sauvegarde le classeur Excel à l'emplacement indiqué.
    Crée les dossiers parents si nécessaire.
    """
    path_dest = Path(chemin_destination)
    path_dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path_dest)
